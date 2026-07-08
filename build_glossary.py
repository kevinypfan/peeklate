"""建立遊戲術語表 glossary.json(英文 → 繁體中文)。

資料來源是兩份社群整理的 The Division 2 資料庫:
- 中文資料庫(公開發佈的 Google Sheets,逐分頁抓 CSV)
- 英文資料庫(Google Sheets 匯出 xlsx)

兩邊沒有共同鍵可以直接 join,所以流程是:
1. 下載並快取到 data/(已存在就跳過,--refresh 強制重抓)
2. 從兩邊各抽出同類別的名稱清單(奇特物品、天賦、套裝、技能…)
3. 每個類別丟給 Gemini 做中英配對(它熟悉官方譯名)
4. 本地驗證: 每組 en/zh 必須逐字存在於來源清單,防幻覺
5. 合併輸出 glossary.json

用法:
    uv run --env-file .env python build_glossary.py            # 完整跑
    uv run python build_glossary.py --extract-only             # 只看抽取結果,不打 API
    uv run --env-file .env python build_glossary.py --refresh  # 強制重新下載
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

DATA_DIR = Path(__file__).parent / "data"
GLOSSARY_PATH = Path(__file__).parent / "glossary.json"

ZH_PUB_BASE = (
    "https://docs.google.com/spreadsheets/u/1/d/e/"
    "2PACX-1vQHp4_A7I8a9PrqU3fe4M3BOxGedQcToT9DptKGOePjbSh9igBt3g04uOVpmo1n5ClqnbDUqQeM7RUY/pub"
)
# 中文資料庫各分頁的 gid(從 pubhtml 頁面取得)
ZH_SHEETS = {
    "exotics": "508181437",
    "weapon_talents": "1503197651",
    "gear_talents": "645885119",
    "brands_gearsets": "915740066",
    "skills": "787637159",
    "attributes": "197681494",
}
EN_XLSX_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1nrPBmOrtpkEW1j5fbcRT7L-AXgsGOqMqxXoVtopsiGM/export?format=xlsx"
)

_CJK = re.compile(r"[一-鿿]")
_RELEASE_LABELS = {"Base Game", "Warlords of New York", "Battle for Brooklyn"}


def download(refresh: bool) -> None:
    import requests

    DATA_DIR.mkdir(exist_ok=True)
    (DATA_DIR / "zh").mkdir(exist_ok=True)
    for name, gid in ZH_SHEETS.items():
        path = DATA_DIR / "zh" / f"{name}.csv"
        if path.exists() and not refresh:
            continue
        print(f"下載中文分頁 {name}…")
        r = requests.get(
            f"{ZH_PUB_BASE}?gid={gid}&single=true&output=csv", timeout=60
        )
        r.raise_for_status()
        path.write_bytes(r.content)
    en_path = DATA_DIR / "en_db.xlsx"
    if not en_path.exists() or refresh:
        print("下載英文資料庫 xlsx…")
        r = requests.get(EN_XLSX_URL, timeout=120)
        r.raise_for_status()
        en_path.write_bytes(r.content)


# ── 抽取:中文 CSV ──────────────────────────────────────────────


def _read_csv(name: str) -> list[list[str]]:
    path = DATA_DIR / "zh" / f"{name}.csv"
    with path.open(encoding="utf-8-sig", newline="") as f:
        return [[c.strip() for c in row] for row in csv.reader(f)]


def _col_by_header(rows: list[list[str]], header: str, spread: int = 0) -> list[str]:
    """找到含指定標題的欄,回傳該欄底下所有非空值(去重、保序)。

    spread: 額外多收右邊幾欄(合併儲存格會讓值落到標題欄旁邊)。
    """
    for r_idx, row in enumerate(rows[:5]):
        for c_idx, cell in enumerate(row):
            if cell == header:
                seen: dict[str, None] = {}
                for row2 in rows[r_idx + 1 :]:
                    for c in range(c_idx, c_idx + spread + 1):
                        if len(row2) > c and row2[c]:
                            seen.setdefault(row2[c], None)
                return list(seen)
    return []


def _zh_clean(vals: list[str]) -> list[str]:
    """保留含中日韓字元、長度合理的詞,去掉說明句(過長或含句號)。

    名稱裡的換行是儲存格排版,接回成一個詞(如「土狼」\\n的面具)。
    """
    out = []
    for v in vals:
        v = "".join(v.split())
        if not v or not _CJK.search(v) or len(v) > 20:
            continue
        if "。" in v or "," in v or "，" in v or re.match(r"^\d", v):
            continue
        out.append(v)
    return out


def extract_zh() -> dict[str, list[str]]:
    cats: dict[str, list[str]] = {}
    cats["exotics"] = _zh_clean(_col_by_header(_read_csv("exotics"), "名稱"))
    cats["exotic_talents"] = _zh_clean(_col_by_header(_read_csv("exotics"), "天賦"))
    cats["weapon_talents"] = _zh_clean(
        _col_by_header(_read_csv("weapon_talents"), "名稱")
    )
    gear_rows = _read_csv("gear_talents")
    cats["gear_talents"] = _zh_clean(_col_by_header(gear_rows, "名稱"))
    # 具名完美欄有合併儲存格,值會落在標題欄或右一欄
    cats["named_gear"] = _zh_clean(_col_by_header(gear_rows, "具名完美", spread=1))
    # 品牌/裝備組分頁: 名稱都在第一欄,品牌與套裝混在一起
    bg_rows = _read_csv("brands_gearsets")
    col0: dict[str, None] = {}
    for row in bg_rows:
        if row and row[0] and row[0] not in ("品牌", "套裝", "裝備組"):
            col0.setdefault(row[0], None)
    cats["brands_gearsets"] = [v for v in col0 if _CJK.search(v) and len(v) <= 20]
    sk_rows = _read_csv("skills")
    cats["skills"] = _zh_clean(_col_by_header(sk_rows, "技能")) + _zh_clean(
        _col_by_header(sk_rows, "分支")
    )
    # 屬性數值分頁沒有標題列,屬性名在第 2 欄
    attr_rows = _read_csv("attributes")
    cats["attributes"] = _zh_clean(
        list(dict.fromkeys(r[1] for r in attr_rows if len(r) > 1 and r[1]))
    )
    return cats


# ── 抽取:英文 xlsx ─────────────────────────────────────────────


def _en_clean(vals: list[str]) -> list[str]:
    out: dict[str, None] = {}
    for v in vals:
        v = " ".join(str(v).split())  # 名稱內的換行接回一個空格
        if (
            not v
            or len(v) > 40
            or v in _RELEASE_LABELS
            or v.startswith("---")
            or _CJK.search(v)
            or not re.search(r"[A-Za-z]{2,}", v)
            or re.search(r"[.!?]\s", v)  # 說明句
            or (v.isupper() and len(v) > 4)  # 槍種區段標題(RIFLES 等)
        ):
            continue
        out.setdefault(v, None)
    return list(out)


def _sheet_col(ws, header: str, header_rows: int = 3) -> list[str]:
    """在前幾列找到以指定標題開頭的欄,回傳底下所有非空字串。"""
    rows = list(ws.iter_rows(max_row=header_rows, values_only=True))
    for r_idx, row in enumerate(rows):
        for c_idx, cell in enumerate(row):
            if cell is not None and str(cell).strip().startswith(header):
                vals = []
                for row2 in ws.iter_rows(min_row=r_idx + 2, values_only=True):
                    if len(row2) > c_idx and row2[c_idx] is not None:
                        vals.append(str(row2[c_idx]))
                return vals
    return []


def extract_en() -> dict[str, list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(DATA_DIR / "en_db.xlsx", read_only=True, data_only=True)
    cats: dict[str, list[str]] = {}
    cats["exotics"] = _en_clean(
        _sheet_col(wb["Weapons Named + Exotics"], "Name")
        + _sheet_col(wb["Gear Named + Exotics"], "Name")
    )
    # 奇特天賦欄是「天賦名\n說明文字」,取第一行(標題用前綴比對,完整字樣過長)
    talent_cells = _sheet_col(wb["Weapons Named + Exotics"], "Talent (")
    cats["exotic_talents"] = _en_clean(
        [c.split("\n", 1)[0] for c in talent_cells]
    )
    cats["weapon_talents"] = _en_clean(_sheet_col(wb["Weapon Talents"], "ALL WEAPONS"))
    cats["gear_talents"] = _en_clean(_sheet_col(wb["Gear Talents"], "Talent"))
    cats["named_gear"] = _en_clean(_sheet_col(wb["Gear Named + Exotics"], "Name"))
    cats["brands_gearsets"] = _en_clean(
        _sheet_col(wb["Brandsets"], "Brand") + _sheet_col(wb["Gearsets"], "Name")
    )
    # 技能清單: Quick Links 欄是「▶ 技能名」,變體在同列右側各欄
    skills: list[str] = []
    for row in wb["Skill List"].iter_rows(max_col=10, values_only=True):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if len(cells) > 2 and cells[2].startswith("▶"):
            skills.append(cells[2].lstrip("▶ "))
            skills.extend(c for c in cells[3:] if c)
    cats["skills"] = _en_clean(skills)
    # 屬性欄有「槍種:\n屬性名」前綴形式,取最後一行
    attr_cells = _sheet_col(wb["Attribute Info"], "Attribute")
    cats["attributes"] = _en_clean([str(c).rsplit("\n", 1)[-1] for c in attr_cells])
    wb.close()
    return cats


# ── Gemini 配對 ────────────────────────────────────────────────

CATEGORY_DESC = {
    "exotics": "exotic/named weapon and gear item names",
    "exotic_talents": "exotic item talent names",
    "weapon_talents": "weapon talent names",
    "gear_talents": "gear (chest/backpack) talent names",
    "named_gear": "named gear item names",
    "brands_gearsets": "gear brand and gear-set names",
    "skills": "skill and skill-variant names",
    "attributes": "gear/weapon attribute stat names",
}


# 建表專用模型: 一次性工作,用比 flash-lite 高一級的 flash(配對更準,容量池也不同)
BUILD_MODEL = "google:gemini-2.5-flash"


def match_category(cat: str, en_list: list[str], zh_list: list[str]) -> dict[str, str]:
    from pydantic import BaseModel
    from pydantic_ai import Agent

    class TermPair(BaseModel):
        en: str
        zh: str

    agent = Agent(
        BUILD_MODEL,
        instructions=(
            "You are matching game terminology from Tom Clancy's The Division 2. "
            "You get two lists of the same category of terms: English originals and "
            "official Traditional Chinese localizations. Pair each Chinese term with "
            "its English original. Copy both strings verbatim from the lists. "
            "Only output pairs you are confident about; omit terms with no match."
        ),
        output_type=list[TermPair],
    )
    prompt = (
        f"Category: {CATEGORY_DESC[cat]}\n\n"
        f"English terms:\n" + "\n".join(en_list) + "\n\n"
        f"Traditional Chinese terms:\n" + "\n".join(zh_list)
    )
    import time

    from pydantic_ai.exceptions import ModelHTTPError

    for attempt in range(4):
        try:
            result = agent.run_sync(prompt)
            break
        except ModelHTTPError as e:
            if e.status_code in (429, 500, 502, 503, 504) and attempt < 3:
                print(f"  ({cat}: HTTP {e.status_code},{2 * 2**attempt} 秒後重試)")
                time.sleep(2 * 2**attempt)
                continue
            raise
    en_set, zh_set = set(en_list), set(zh_list)
    pairs = {}
    dropped = 0
    for p in result.output:
        if p.en in en_set and p.zh in zh_set:
            pairs[p.en] = p.zh
        else:
            dropped += 1
    if dropped:
        print(f"  ({cat}: 丟棄 {dropped} 組不在來源清單中的配對)")
    return pairs


def main() -> None:
    parser = argparse.ArgumentParser(description="建立 glossary.json 術語表")
    parser.add_argument("--refresh", action="store_true", help="強制重新下載資料庫")
    parser.add_argument(
        "--extract-only", action="store_true", help="只印抽取結果,不呼叫 API"
    )
    args = parser.parse_args()

    download(args.refresh)
    zh_cats = extract_zh()
    en_cats = extract_en()

    for cat in CATEGORY_DESC:
        print(f"{cat}: 英 {len(en_cats.get(cat, []))} 條, 中 {len(zh_cats.get(cat, []))} 條")
    if args.extract_only:
        for cat in CATEGORY_DESC:
            print(f"\n=== {cat} ===")
            print("EN:", en_cats.get(cat, [])[:100])
            print("ZH:", zh_cats.get(cat, [])[:100])
        return

    glossary: dict[str, str] = {}
    for cat in CATEGORY_DESC:
        en_list, zh_list = en_cats.get(cat, []), zh_cats.get(cat, [])
        if not en_list or not zh_list:
            print(f"{cat}: 清單為空,跳過")
            continue
        pairs = match_category(cat, en_list, zh_list)
        print(f"{cat}: 配出 {len(pairs)} 組")
        glossary.update(pairs)

    GLOSSARY_PATH.write_text(
        json.dumps(glossary, ensure_ascii=False, indent=1, sort_keys=True),
        encoding="utf-8",
    )
    print(f"\n共 {len(glossary)} 條 → {GLOSSARY_PATH}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
